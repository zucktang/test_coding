import openpyxl
import json

def read_campaign_data(sheet):
    header_row = None
    for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
        if 'campaign' in str(row[0]).lower():
            header_row = row
            break
        

    if header_row is not None:
        data = []
        for row in sheet.iter_rows(min_row=header_row[0] + 1, values_only=True):
            row_dict = dict(zip(header_row, row))
            data.append(row_dict)
        return {"data": data}
    return None

def read_interest_rate_data(sheet):
    header = [cell.value for cell in sheet[1]]
    if header[0].lower() == 'interest rate':
        data = [header] + [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2, values_only=True)]
        return {"interest": data}
    return None

def read_excel(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    result_dict = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        campaign_data = read_campaign_data(sheet)
        if campaign_data is not None:
            result_dict[sheet_name] = campaign_data
        else:
            interest_data = read_interest_rate_data(sheet)
            if interest_data is not None:
                result_dict[sheet_name] = interest_data

    return result_dict

file_path = 'hw4.xlsx'
result_dict = read_excel(file_path)

json_result = json.dumps(result_dict, default=str, indent=2, ensure_ascii=False)
print(json_result)
