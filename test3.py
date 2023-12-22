import openpyxl
from datetime import datetime

def read_excel(file_name):
    wb = openpyxl.load_workbook(file_name, data_only=True)
    yellow_color = 'FFFFFF00'
    red_color = 'FFFF0000'
    
    result = {}
    current_campaign = None

    for sheet in wb:
        for row in sheet.iter_rows(min_row=2):  # Assuming header is in the first row
            campaign_entry = {}
            interest_rate_entry = []

            for cell in row:
                col_letter = openpyxl.utils.get_column_letter(cell.column)

                if cell.fill.fgColor.rgb == yellow_color:
                    
                    current_campaign = cell.value
                    result[current_campaign] = {"data": [], "interest": []}
                elif cell.fill.fgColor.rgb == red_color:
                    interest_rate_entry.append(cell.value)
                else:
                    if col_letter == 'B':
                        campaign_entry["effdate"] = cell.value.strftime('%Y/%m/%d') if isinstance(cell.value, datetime) else cell.value
                    elif col_letter == 'C':
                        campaign_entry["filename"] = cell.value
                    elif col_letter == 'F':
                        campaign_entry["remark"] = cell.value
                import pdb;pdb.set_trace()
            if current_campaign and campaign_entry:
                result[current_campaign]["data"].append(campaign_entry)

            if current_campaign and interest_rate_entry:
                result[current_campaign]["interest"].append(interest_rate_entry)

    return result

# Example usage
file_name = 'hw4.xlsx'
result = read_excel(file_name)
# print(result)
