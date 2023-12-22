import openpyxl


def get_next_row_cell(cell_coordinate, i=1):
    col_letter = cell_coordinate[0]
    col_number = cell_coordinate[1:]
    new_col_number = int(col_number) + i
    new_coordinate = f"{col_letter}{new_col_number}"
    return new_coordinate

def get_next_col_cell(cell_coordinate, i=1):
    col_letter = cell_coordinate[0]
    new_col_letter = chr(ord(col_letter) + i)
    new_coordinate = f"{new_col_letter}{cell_coordinate[1:]}"
    return new_coordinate
    
def read_excel(file_name):
    wb = openpyxl.load_workbook(file_name, data_only=True)
    yellow_color = 'FFFFFF00'
    red_color = 'FFFF0000'
    result = {}
    for sheet in wb:
        campaign_cols = {}
        yellow_background_cols = []
        interest_rate_cols = {}
        campaign_values = ['filename', 'effdate', 'remark']
        campaign_name = ''
        for col in sheet.iter_cols():
            for cell in col:
                if cell.fill.fgColor.rgb == yellow_color:
                    if 'campaign' in str(cell.value).lower():
                        campaign_name = cell.value
                    campaign_cols[cell.coordinate] = cell.value
                if cell.value in campaign_values:
                    yellow_background_cols.append(cell.coordinate) # [{'effdate': 'B3'}, {'filename': 'C3'}, {'remark': 'F3'}]
                if cell.font.color.rgb == red_color:
                    interest_rate_cols[cell.coordinate] = cell.value
                 
        data_dict = {'effdate':None,
                'filename':None,
                'remark': None,
                }
        data_list = []
        
        for col in yellow_background_cols:
            column = campaign_cols[col]
            next_row = get_next_row_cell(col)
            
            if next_row in campaign_cols:
                value = campaign_cols[next_row]
                data_dict[column] = value if column != 'effdate' else value.strftime("%Y/%m/%d")
                data_list.append(data_dict)
        interest_list = []
        for index, (key, value) in enumerate(interest_rate_cols.items()):
            data_in_rows = []
            for i in range(3):
                next_col = get_next_col_cell(key,i)
                if next_col in interest_rate_cols:
                    data_in_rows.append(interest_rate_cols[next_col])
            interest_list.append(data_in_rows)
            
            if index == 2:
                break
def main():
    file_name = 'hw4.xlsx'
    result = read_excel(file_name)
    workbook = openpyxl.load_workbook('hw4.xlsx')
    
    expected = {
            "campaign1": {
                "data": [
                {"effdate": "2022/10/10", "filename": "x.txt", "remark": "w"},
                {"effdate": "2022/10/11", "filename": "y.txt", "remark": "kk"},
                {"effdate": "2022/10/11", "filename": "xx", "remark": "kk"},
                {"effdate": "2022/12/11", "filename": "12", "remark": "79"}
                ],
                "interest": [
                ["interest rate", 12, 24, 12, 0.01, 0.02, 0.02, 0.03],
                [2000, 0.02, 0.03],
                [3000, 0.02, 0.03],
                [4000, 0.02, 0.03]
                ]
            },
            "testCampaign": {
                "data": [
                {"effdate": "2022/10/10", "filename": "c2", "remark": "col"},
                {"effdate": "2022/10/11", "filename": "", "remark": "cel"}
                ],
                "interest": [
                ["interest rate", 12, 24, 12, 0.01, 0.02, 3, 0.02, 0.03]
                ]
            }
            }


if __name__ == "__main__":
    main()

